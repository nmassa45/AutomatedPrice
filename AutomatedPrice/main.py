"""This is the main python file for this price change program.

This script should complete the process for updating the prices for the master sheet filled with products. It requires
the update prices excel file and the master excel file both filled with prices. This script requires that openpyxl be
installed, itertools, and it uses PatternFill from openpyxl.

Make sure that the two files you are using are downloaded as .xlsx extensions.

Also make sure to remember that the sheets on each worksheet that you use should always be renamed to "info". This
is the universal standard in this program for all sheets to be named this, and is how they get called directly when
need be.

"""
import locale
import time
from itertools import chain
import openpyxl
from openpyxl.styles import PatternFill


def collect_information(workbook_name, sheet_name, columns, starting_row_number, finishing_row_number):
    """
    Collect information from excel file

    This function should be used to collect data from a given excel file. In this project we are normally trying
    to collect the SKU number of a product and the respectable price.

    :param workbook_name: This represents the workbook name of the excel file we are collecting.
    :param sheet_name: This represents the name of the sheet, usually just change it to "info"
    :param columns: An array containing the letters of the SKU number and the price.
    :param starting_row_number: Represents the starting position.
    :param finishing_row_number: Represents the ending position.
    :return: This will return a list where each element is a list containing the SKU number and the price.
    """
    master_list = []
    wb = openpyxl.load_workbook(workbook_name)
    sheet = wb[sheet_name]

    for i in range(starting_row_number, finishing_row_number + 1):
        row_properties = []
        row_string = str(i)
        for j in range(len(columns)):  # zero is sku, one is price
            column_letter = columns[j]
            value = sheet[column_letter + row_string].value
            if j == 0:
                value = str(sheet[column_letter + row_string].value).strip().upper()
            elif j == 1:
                floatPrice = float(value)
                roundPrice = round(floatPrice, 2)
                value = str(roundPrice)
                print("This was the rounded Price: " + value)
            row_properties.append(value)

        master_list.append(row_properties)
        print("Cell information was collected: " + str(i))
    print("finished")
    return master_list


def check_which_products_exist(first_list, second_list):
    """
    Find matches between two lists.

    This function takes one list and compares it with the second list. In our case the first case is the price increases
    and the second list is the master list.
    :param first_list: Represents the first list filled with elements that contain a SKU number and a price.
    :param second_list: Represents the list we are comparing to.
    :return: This will return a list of lists where each element has a SKU number and a price, but specifically it is
    only the ones that had a match.
    """
    matched_ids = []
    matched_count = 0
    for i in first_list:
        product_id = i[0]
        print("This is the product id: " + product_id)
        is_found = product_id in chain(*second_list)
        if is_found:
            print("Match")
            matched_count = matched_count + 1
            matched_ids.append(i)
    print("Matched: " + str(matched_count) + " out of: " + str(len(first_list)))
    return matched_ids


def update_price(matched_list, workbook_name, start_row, last_row):
    """
    Update the price in the master sheet.

    Function is used to update the master sheet with the matched_list element. This element is a list that contains
    matched SKU numbers and the new price that needs to be added. This function should not return anything, should
    simply update the products in the master sheet.

    :param matched_list: Represents a list of lists with matched ID's.
    :param workbook_name: Represents the name of the workbook we are trying to update.
    :param start_row: Represents where we are starting.
    :param last_row: Represents where we are ending.
    """
    print("we are inside this update function")
    wb = openpyxl.load_workbook(workbook_name)
    ws = wb.active
    green_fill = PatternFill(start_color='0000FF00',
                             end_color='0000FF00',
                             fill_type='solid')
    # Be careful when updating the master sheet because it might be the wrong column.
    for i in matched_list:  # The i's represent the product properties
        product_id = i[0]
        results = False
        while not results:
            for row in ws.iter_rows(None, start_row, last_row):  # If error occurs with iter_rows, check parameters in worksheet.py
                if str(row[1].value).strip().upper() == product_id:
                    print("ID: " + str(row[1].value) + " was found and the price is: " + str(row[4].value))
                    if str(row[4].value).strip() == '0':  # Checking the master list if the price is zero
                        print("We do not work with this product!")
                        results = True
                    else:
                        row[4].value = i[1]
                        print("Updated the price with: " + str(i[1]))
                        highlight_row(row, green_fill)
                        results = True
    wb.save(workbook_name)
    print("Saved the " + workbook_name + " workbook!")


def high_light_price_increase(matched_list, workbook_name, start_row, last_row):
    """
    Highlighting rows in the price increase sheet.

    Function for highlighting the rows in the price increase sheet. For example a row that has a match in the master
    sheet should be highlighted green, and rows that are not found should be highlighted yellow. This function should
    not return anything. It should just highlight rows in the excel file.

    :param matched_list: Represents a list of matched ID's.
    :param workbook_name: Represents the workbook name we are looking at.
    :param start_row: Represents where we are starting.
    :param last_row: Represents where we are ending.
    """
    wb = openpyxl.load_workbook(workbook_name)
    ws = wb.active
    green_fill = PatternFill(start_color='0000FF00',
                             end_color='0000FF00',
                             fill_type='solid')
    yellow_fill = PatternFill(start_color='00FFFF00',
                              end_color='00FFFF00',
                              fill_type='solid')

    for row in ws.iter_rows(None, start_row, last_row):  # Was updated to pass None into iter_rows to prevent crash 6-22-21
        row_id = str(row[0].value).strip()  # This specifies column A in the row iteration
        print("We are in this loop for highlighting")
        is_found = row_id in chain(*matched_list)
        if is_found:
            print("The ID: " + row_id + " was found!")
            highlight_row(row, green_fill)
        else:
            print("The ID: " + row_id + " was not found!")
            highlight_row(row, yellow_fill)
    wb.save(workbook_name)
    print("Saved the workbook!")


def highlight_row(row, pattern):
    """
    Highlight row

    Helper method for highlighting each cell in a specific row. This function should not return anything. It should
    just change the background for each cell in the specific row.

    :param row: Represents the row we are looking at.
    :param pattern: Represents the openpxl pattern style we wish to do to each cell in the row.
    """
    for cell in row:
        cell.fill = pattern


def import_excel_price_increase(properties, start, sheet, workbook, workbook_name):
    """
    Really its just a method to import into excel

    This method focuses on importing the data into an excel sheet. This data focuses more on the products on our
    website which have a higher price when compared to a competitor.

    :param properties: Represents a list of lists, where each list contains properties about a product.
    :param start: This is the position to start importing in the excel file.
    :param sheet: This is the actual sheet of the workbook.
    :param workbook: This is the actual workbook.
    :param workbook_name: This is the name of the workbook, which is needed to save the file.
    """
    columns = ['A', 'B', 'C', 'D']
    print("This is where we need to start: " + str(start))
    for i in range(len(properties)):
        row_pointer = start + i
        print("position: " + str(row_pointer))
        specific_properties = properties[i]
        for j in range(len(specific_properties)):
            value = specific_properties[j]
            try:
                column = columns[j]
                sheet[column + str(row_pointer)] = value
                time.sleep(.5)
            except IndexError:
                print("This usually is a sign that the site has updated their information!")
        time.sleep(.5)
        print("A row has been completed")
    workbook.save(workbook_name)
    print("The workbook: " + workbook_name + " has been saved!")


def price_update_changes_comparisons(price_increase, price_increase_columns, price_increase_indices,
                                     master, master_columns, master_indices):
    """
    Function for price increases between two files

    This is the main function that collects the information from the price changes excel file, the information from the
    master excel sheet, compares the two lists to find matches on the master excel sheet, highlights the rows in the price
    increase excel sheet that are matches in the master list, and finally updates the price in the master list while
    also highlighting those rows in yellow. Each method requires a couple of parameters to run, information about those can
    be found in each respectable function.

    :param price_increase: This represents the price_increase excel file.
    :param price_increase_columns: This represents a list containing the columns for the sku number and the price.
    :param price_increase_indices: This represents a list containing indices for start and finish in the excel file.
    Usually start at 2 because the first row are titles, and the last row is exclusive, but this gets delt with in the
    'collect_information function'.
    :param master: This represents the master list from PSC filled with products.
    :param master_columns: This represents the a list containing the columns for the sku number and the price
    Ex ['A, D'].
    :param master_indices: This represents a list containing the indices for start and finish in the excel file.
    """

    price_changes = collect_information(price_increase, 'info', price_increase_columns,
                                        price_increase_indices[0], price_increase_indices[1])
    compare_information = collect_information(master, 'info', master_columns, master_indices[0], master_indices[1])
    matched = check_which_products_exist(price_changes, compare_information)
    high_light_price_increase(matched, price_increase, price_increase_indices[0], price_increase_indices[1])
    print("We are going to update the master sheet now")
    update_price(matched, master, master_indices[0], master_indices[1])


def compare_Scrape_Verus_Master(scrape_fileName, scrape_sheetName, scrape_columns, scrape_start, scrape_end,
                                master_fileName, new_compare_file, new_compare_sheet):
    """
    Function for getting a list of products with higher prices on the master sheet.

    Function for creating a new excel file, which would contain only the products that were found to have
    a higher price on the master sheet versus the scraped website. Therefore, these products should have a price
    decrease.

    :param scrape_fileName: This will represent the file name with all of the scraped products.
    :param scrape_sheetName: This will represent the sheet name with all of the scraped products.
    :param scrape_columns:  This will represent an array of columns representing the price and the SKU number.
    :param scrape_start: This will represent the starting position in the excel scrape file.
    :param scrape_end: This will represent the ending position in the excel scrape file.
    :param master_fileName:  This will represent the name of the master sheet containing all of the products.
    :param new_compare_file:  This will represent the name of the file which will contain the products with higher
    prices found on the master sheet when compared to the scrape.
    :param new_compare_sheet: This will represent the sheet name of the compare file.
    :return: This will not return anything, will fill the compare file, which was previously created.
    """
    comparisons = []
    scrape_collect = collect_information(scrape_fileName, scrape_sheetName, scrape_columns, scrape_start, scrape_end)
    wb = openpyxl.load_workbook(master_fileName)  # Focusing on the master list
    ws = wb.active
    # We now know that the two collections are of size two in the form [price, ID]
    for i in scrape_collect:  # The i represents products in our scrape_collect.
        product_id = i[1][3:]  # product in scrape # After this comparison, get rid of this because it will be fixed.
        print("scrape id: " + product_id)
        local_comparison = []
        local_comparison.append(product_id)
        for row in ws.iter_rows(2, 1014):  # This is looping through the master sheet.
            print("This is the id value: " + str(row[0].value))
            if str(row[0].value).strip().upper() == product_id:  # compares master sheet to scrape
                print("ID: " + str(row[0].value) + " was found and the price in master sheet is: " + str(row[1].value))
                # compares the row price in master sheet to row price in scrape
                comparePrice = row[1].value
                scrape_price = i[0].replace('$', '').strip()
                if str(comparePrice) == "*overflow*":
                    print("This price was not available!")
                    local_comparison.append("not available")
                    break
                if scrape_price == "SOLD OUT":
                    print("This price was not available!")
                    local_comparison.append("not available")
                    break
                locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
                print(comparePrice)
                print(locale.atof(scrape_price))
                if comparePrice > locale.atof(scrape_price):
                    print("In the most recent scrape this price decreased")
                    local_comparison.append("Sandblasting HIGHER")  # Depends on what store
                    local_comparison.append(comparePrice)
                    local_comparison.append(locale.atof(scrape_price))
                    break
            else:
                print("This is a case that we do not care about")
        if len(local_comparison) <= 2:
            print("We will not be adding this comparison to the list!")
        else:
            comparisons.append(local_comparison)
        print("Moving onto the next product in the scrape")
    wb = openpyxl.load_workbook(new_compare_file)
    sheet = wb[new_compare_sheet]
    import_excel_price_increase(comparisons, 2, sheet, wb, new_compare_file)


if __name__ == '__main__':
    price_update_changes_comparisons("PriceIncreases/BigC Update Chevron June '21 Increase.xlsx", ['A', 'D'], [2, 132],
                                     "MasterSheets/products-2021-06-22.xlsx", ['B', 'E'], [2, 2323])
    # print("Just wanted to test something again!")
